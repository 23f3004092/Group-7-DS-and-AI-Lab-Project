"""CACP-based cost service: live MSP from cacp.da.gov.in and per-hectare cost of
cultivation from the DES consolidated workbook (the official cost estimates used
by CACP). Both follow the mandi/weather pattern: live fetch with cache + static
fallback constants when the source is unreachable."""
import io
import time
import asyncio
from typing import Dict, Optional, Any, List, Tuple
import httpx
from ..config import settings

CROP_SHEETS = {
    "wheat": "Wheat",
    "rice": "Paddy",
    "paddy": "Paddy",
    "maize": "Maize",
    "mustard": "R&M",
    "sarson": "R&M",
    "potato": "Potato",
    "sugarcane": "Sugarcane",
}

# Cost A2+FL per hectare (INR/ha) per crop — fallback when the DES workbook is unreachable
COST_FALLBACK = {
    "wheat": 32000.0,
    "rice": 38000.0,
    "paddy": 38000.0,
    "maize": 25000.0,
    "mustard": 25000.0,
    "sarson": 25000.0,
    "potato": 45000.0,
    "sugarcane": 60000.0,
}

# MSP (INR/quintal) fallback when the CACP API is unreachable
MSP_FALLBACK = {
    "wheat": 2275.0,
    "rice": 2183.0,
    "paddy": 2183.0,
    "maize": 2090.0,
    "mustard": 5650.0,
    "sarson": 5650.0,
    "potato": 1080.0,
    "sugarcane": 3700.0,
}


class CostService:
    def __init__(self):
        self._cache: Dict[str, Tuple[float, Any]] = {}
        self._cache_ttl = settings.CACP_CACHE_TTL
        self._cost_data = None

    def _get_cached(self, key: str) -> Optional[Any]:
        entry = self._cache.get(key)
        if entry and time.time() - entry[0] < self._cache_ttl:
            return entry[1]
        return None

    def _set_cached(self, key: str, data: Any):
        self._cache[key] = (time.time(), data)

    def normalize_crop(self, crop: str) -> str:
        return crop.lower().strip()

    def fallback_cost(self, crop: str) -> float:
        return COST_FALLBACK.get(self.normalize_crop(crop), 20000.0)

    def fallback_msp(self, crop: str) -> float:
        return MSP_FALLBACK.get(self.normalize_crop(crop), 1800.0)

    # --- Live MSP from CACP ---

    async def get_msp(self, crop: str) -> Optional[float]:
        """Latest CACP-recommended/fixed MSP (INR/quintal) for the crop."""
        crop_n = self.normalize_crop(crop)
        key = f"msp|{crop_n}"
        cached = self._get_cached(key)
        if cached is not None:
            return cached

        msp = None
        try:
            records = await self._fetch_msp_records()
            match_key = "paddy" if crop_n in ("rice", "paddy") else crop_n
            for r in records:
                commodity = str(r.get("commodityname", "")).lower()
                if commodity.startswith(match_key) or match_key in commodity:
                    try:
                        msp = float(r.get("fixed_price") or r.get("reco_price"))
                    except (TypeError, ValueError):
                        continue
                    break
        except Exception as e:
            print(f"CACP MSP fetch failed: {e}. Using static MSP fallback.")

        if msp is None:
            msp = self.fallback_msp(crop_n)
        self._set_cached(key, msp)
        return msp

    async def _fetch_msp_records(self) -> List[Dict[str, Any]]:
        cached = self._get_cached("msp|records")
        if cached is not None:
            return cached

        # cacp.da.gov.in sits behind Google's frontend and intermittently 403s
        # non-browser or rapid requests. Send a browser-like UA and retry the
        # transient 403/429/5xx errors before giving up.
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "KHTML, like Gecko) Chrome/126.0 Safari/537.36"
            ),
            "Accept": "application/json",
        }
        async with httpx.AsyncClient(timeout=30.0, headers=headers, follow_redirects=True) as client:
            for attempt in range(3):
                try:
                    resp = await client.get(settings.CACP_MSP_URL)
                    if resp.status_code in (403, 429) or resp.status_code >= 500:
                        if attempt < 2:
                            await asyncio.sleep(1.5 * (attempt + 1))
                            continue
                        resp.raise_for_status()
                    resp.raise_for_status()
                    records = resp.json()
                    break
                except (httpx.TransportError, httpx.HTTPStatusError) as e:
                    if attempt >= 2:
                        raise
                    await asyncio.sleep(1.5 * (attempt + 1))
        self._set_cached("msp|records", records)
        return records

    # --- Cost of cultivation per hectare from DES workbook ---

    async def get_cost_per_ha(self, crop: str, state: str = "Uttar Pradesh") -> Optional[float]:
        """CACP/DES cost of cultivation (INR/ha, Cost A2+FL) for crop x state."""
        crop_n = self.normalize_crop(crop)
        key = f"cost|{crop_n}|{state.lower()}"
        cached = self._get_cached(key)
        if cached is not None:
            return cached

        cost = None
        try:
            cost = await self._parse_cost(crop_n, state)
        except Exception as e:
            print(f"CACP cost data fetch failed: {e}. Using static cost fallback.")

        if cost is None:
            cost = self.fallback_cost(crop_n)
        self._set_cached(key, cost)
        return cost

    async def _parse_cost(self, crop_n: str, state: str) -> Optional[float]:
        sheet = CROP_SHEETS.get(crop_n)
        if sheet is None:
            return None

        book = await self._load_workbook()
        if book is None or sheet not in book.sheetnames:
            return None
        ws = book[sheet]

        header_row = None
        states = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and str(row[0]).strip().lower() == "sl no":
                header_row = row
                for ci, c in enumerate(row):
                    if ci >= 3 and c and not str(c).strip().isdigit():
                        states[ci] = str(c).strip()
                break
        if header_row is None or not states:
            return None

        target = next((ci for ci, s in states.items() if s.lower() == state.lower()), None)
        if target is None:
            return None

        for row in ws.iter_rows(values_only=True):
            code = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if code == "A2+FL" and row[target] is not None:
                return float(row[target])
        return None

    async def _load_workbook(self):
        if self._cost_data is not None:
            return self._cost_data
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            resp = await client.get(settings.COST_DATA_URL)
            resp.raise_for_status()
        import openpyxl
        self._cost_data = openpyxl.load_workbook(
            io.BytesIO(resp.content), read_only=True, data_only=True
        )
        return self._cost_data


cost_service = CostService()
